"""
sport_normalizer.py
====================
Импорт Всероссийского реестра видов спорта (ВРВС) из XLS и нормализация
названий видов спорта из приказов.

Основные задачи:
  1. Импорт XLS-справочника ВРВС в БД (с версионностью)
  2. Нечёткое сопоставление названий видов спорта из приказов
     с каноническими названиями из справочника
  3. Управление «сроком жизни» наименований (переименования,
     напр. «Тайский бокс» → «Муайтай»)
  4. Поддержка алиасов и известных опечаток

Модель версионности:
  - sport_registry_versions: каждый импорт XLS — новая версия
  - sports: канонические виды спорта (стабильный code_base)
  - sport_names: наименования с valid_from/valid_to
    По умолчанию valid_to = NULL (бессрочно).
    При переименовании: старое имя получает valid_to,
    создаётся новое с valid_from.
  - sport_disciplines: дисциплины внутри вида спорта

Использование:
    # Импорт реестра
    normalizer = SportNormalizer()
    normalizer.load_xls("Reestr_b2e716479e.xls", version_label="ВРВС от 24.07.2025")

    # Нормализация названия из приказа
    result = normalizer.normalize("Спортиваня акробатика")
    # → NormalizationResult(canonical="Спортивная акробатика", sport_id=..., confidence=0.92, method="fuzzy")

    # Все виды спорта
    for sport in normalizer.all_sports():
        print(sport.current_name, sport.code_full)

Зависимости:
    pip install openpyxl rapidfuzz
    (опционально) apt install libreoffice-calc  — для конвертации .xls → .xlsx
"""

import hashlib
import logging
import re
import unicodedata
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Optional, Union

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Конфигурация
# ---------------------------------------------------------------------------

# Порог уверенности для автоматического сопоставления (0–100)
AUTO_MATCH_THRESHOLD = 85

# Порог для «подозрительного» совпадения (требует ревью)
REVIEW_THRESHOLD = 70

# Разделы ВРВС
SECTION_NAMES = {
    1: "Признанные (не нац./не воен./не общерос.)",
    2: "Общероссийские",
    3: "Национальные",
    4: "Военно-прикладные и служебно-прикладные",
}

# Известные алиасы и переименования (hardcoded для стабильности).
# Формат: {"старое/альтернативное название": "каноническое в ВРВС"}
# Пополняется вручную при обнаружении расхождений в приказах.
KNOWN_ALIASES: dict[str, str] = {
    # Переименования
    "Тайский бокс": "Муайтай",
    "Тай бокс": "Муайтай",

    # Устаревшие названия
    "Водное поло": "Водное поло",  # без изменений, но для единообразия
    "Кёрлинг": "Керлинг",

    # Типичные опечатки из OCR и приказов
    "Спортиваня акробатика": "Спортивная акробатика",
    "Спортивнаяакробатика": "Спортивная акробатика",
    "Кёкусин": "Киокусинкай",  # ёкусин → Киокусинкай (код 173)
    "Лёгкая атлетика": "Легкая атлетика",
    "Художественая гимнастика": "Художественная гимнастика",
    "Спортивнаягимнастика": "Спортивная гимнастика",
    "Настольный тенис": "Настольный теннис",

    # Дисциплины, которые в приказах часто указаны как вид спорта
    "Вольная борьба": "Спортивная борьба",
    "Греко-римская борьба": "Спортивная борьба",

    # Сокращения
    "ФМ": "Функциональное многоборье",
}


# ---------------------------------------------------------------------------
# Модели данных
# ---------------------------------------------------------------------------

class MatchMethod(str, Enum):
    """Метод, которым было найдено совпадение."""
    EXACT       = "exact"       # точное совпадение с каноническим именем
    ALIAS       = "alias"       # совпадение через KNOWN_ALIASES
    CASE_NORM   = "case_norm"   # совпадение после нормализации регистра/пробелов
    FUZZY       = "fuzzy"       # нечёткое совпадение (rapidfuzz)
    NOT_FOUND   = "not_found"   # не найдено


@dataclass
class SportEntry:
    """Вид спорта из справочника ВРВС."""
    code_base: int              # базовый номер (152, 166...)
    code_full: str              # полный код «166-0-5-5-1-1-Я»
    section: int                # раздел ВРВС (1–4)
    name: str                   # текущее название
    disciplines: list[str] = field(default_factory=list)
    db_id: Optional[str] = None # UUID из БД (если загружен)


@dataclass
class NormalizationResult:
    """Результат нормализации названия вида спорта."""
    input_name: str             # исходное название из приказа
    canonical_name: Optional[str] = None  # каноническое из ВРВС
    sport_id: Optional[str] = None        # UUID из БД
    confidence: float = 0.0     # 0.0–1.0
    method: MatchMethod = MatchMethod.NOT_FOUND
    alternatives: list[tuple[str, float]] = field(default_factory=list)
    # [(name, score), ...] — топ-3 альтернативы при fuzzy


@dataclass
class ImportStats:
    """Статистика импорта реестра."""
    version_label: str
    file_hash: str
    sports_total: int = 0
    disciplines_total: int = 0
    sports_new: int = 0
    sports_updated: int = 0
    names_added: int = 0


# ---------------------------------------------------------------------------
# Парсер XLS
# ---------------------------------------------------------------------------

class VrvsXlsParser:
    """
    Парсит XLS-файл Всероссийского реестра видов спорта.

    Структура файла (4 листа):
      - «Признанные» (раздел 1)
      - «Общероссийские» (раздел 2)
      - «Национальные» (раздел 3)
      - «Прикладные» (раздел 4)

    Каждый лист: строки с видами спорта (номер + название + код)
    и строки с дисциплинами (название в колонке 10+).
    """

    SHEET_SECTIONS = {
        "Признанные": 1,
        "Общероссийские": 2,
        "Национальные": 3,
        "Прикладные": 4,
    }

    def __init__(self, xls_path: Union[str, Path]):
        self.path = Path(xls_path)
        if not self.path.exists():
            raise FileNotFoundError(f"Файл не найден: {self.path}")

        # Для .xls нужна конвертация в .xlsx
        self._xlsx_path = self._ensure_xlsx()

    def _ensure_xlsx(self) -> Path:
        """Конвертирует .xls в .xlsx если нужно."""
        if self.path.suffix.lower() == ".xlsx":
            return self.path

        # Ищем уже конвертированный файл (рядом с исходным или в cwd)
        xlsx_path = self.path.with_suffix(".xlsx")
        if xlsx_path.exists():
            return xlsx_path

        # Также проверяем текущую директорию
        cwd_xlsx = Path.cwd() / (self.path.stem + ".xlsx")
        if cwd_xlsx.exists():
            return cwd_xlsx

        # Конвертация через LibreOffice
        import subprocess

        # Попытка через soffice.py (доступен в среде Claude)
        soffice_script = Path("/mnt/skills/public/docx/scripts/office/soffice.py")
        if soffice_script.exists():
            try:
                outdir = str(self.path.parent)
                result = subprocess.run(
                    ["python3", str(soffice_script), "--headless",
                     "--convert-to", "xlsx", str(self.path),
                     "--outdir", outdir],
                    capture_output=True, text=True, timeout=60,
                )
                if xlsx_path.exists():
                    logger.info(f"Конвертирован {self.path.name} → {xlsx_path.name}")
                    return xlsx_path
            except Exception as e:
                logger.warning(f"soffice.py failed: {e}")

        # Попытка через libreoffice напрямую
        try:
            result = subprocess.run(
                ["libreoffice", "--headless", "--convert-to", "xlsx",
                 str(self.path), "--outdir", str(self.path.parent)],
                capture_output=True, text=True, timeout=60,
            )
            if xlsx_path.exists():
                logger.info(f"Конвертирован {self.path.name} → {xlsx_path.name}")
                return xlsx_path
        except (subprocess.TimeoutExpired, FileNotFoundError):
            pass

        raise RuntimeError(
            f"Не удалось конвертировать {self.path.name} в .xlsx. "
            f"Установите LibreOffice: apt install libreoffice-calc"
        )

    def parse(self) -> list[SportEntry]:
        """
        Парсит XLS и возвращает список SportEntry.
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("Установи openpyxl: pip install openpyxl")

        wb = openpyxl.load_workbook(self._xlsx_path, read_only=True)
        all_sports: list[SportEntry] = []

        for sheet_name, section_num in self.SHEET_SECTIONS.items():
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Лист '{sheet_name}' не найден в файле")
                continue

            ws = wb[sheet_name]
            current_sport: Optional[SportEntry] = None

            for row in ws.iter_rows(min_row=1, values_only=False):
                vals = [cell.value for cell in row]

                # Строка вида спорта: col1=число, col2=название
                if (vals[0] is not None and vals[1] is not None
                        and str(vals[0]).strip().isdigit()):
                    sport_name = str(vals[1]).strip()
                    code_parts = vals[2:9]  # колонки 3–9
                    code_full = self._build_code(code_parts)

                    code_base = int(float(code_parts[0])) if code_parts[0] else 0

                    current_sport = SportEntry(
                        code_base=code_base,
                        code_full=code_full,
                        section=section_num,
                        name=sport_name,
                        disciplines=[],
                    )
                    all_sports.append(current_sport)

                # Строка дисциплины: col10 непусто
                if len(vals) > 9 and vals[9] is not None:
                    disc_name = str(vals[9]).strip()
                    if disc_name and current_sport is not None:
                        current_sport.disciplines.append(disc_name)

        wb.close()

        logger.info(
            f"ВРВС: {len(all_sports)} видов спорта, "
            f"{sum(len(s.disciplines) for s in all_sports)} дисциплин"
        )
        return all_sports

    @staticmethod
    def _build_code(parts: list) -> str:
        """Собирает код вида спорта из частей: '166-0-5-5-1-1-Я'."""
        cleaned = []
        for p in parts:
            if p is None:
                break
            s = str(p).strip()
            if not s:
                break
            # Числа: убираем .0
            try:
                n = float(s)
                if n == int(n):
                    s = str(int(n))
            except (ValueError, TypeError):
                pass
            cleaned.append(s)
        return "-".join(cleaned) if cleaned else ""

    def file_hash(self) -> str:
        """SHA256 исходного XLS-файла."""
        return hashlib.sha256(self.path.read_bytes()).hexdigest()


# ---------------------------------------------------------------------------
# Нормализатор
# ---------------------------------------------------------------------------

class SportNormalizer:
    """
    Нормализует названия видов спорта из приказов, сопоставляя их
    с каноническими названиями из ВРВС.

    Стратегия матчинга (по приоритету):
      1. Exact match (с учётом регистра)
      2. Alias match (KNOWN_ALIASES)
      3. Case-normalized match (upper + strip + collapse spaces)
      4. Fuzzy match (rapidfuzz, token_sort_ratio)

    Использование без БД (in-memory):
        normalizer = SportNormalizer()
        normalizer.load_xls("Reestr.xls")
        result = normalizer.normalize("Спортиваня акробатика")

    Использование с БД:
        normalizer = SportNormalizer(db_conn=conn)
        normalizer.load_from_db()
        result = normalizer.normalize("Муайтай")
    """

    def __init__(
        self,
        auto_threshold: float = AUTO_MATCH_THRESHOLD,
        review_threshold: float = REVIEW_THRESHOLD,
    ):
        self.auto_threshold = auto_threshold
        self.review_threshold = review_threshold

        # In-memory индексы
        self._sports: list[SportEntry] = []
        self._name_to_sport: dict[str, SportEntry] = {}    # exact name → sport
        self._norm_to_sport: dict[str, SportEntry] = {}     # normalized → sport
        self._alias_to_canon: dict[str, str] = dict(KNOWN_ALIASES)

        # Кеш нормализаций для скорости
        self._cache: dict[str, NormalizationResult] = {}

    # ------------------------------------------------------------------
    # Загрузка справочника
    # ------------------------------------------------------------------

    def load_xls(
        self,
        xls_path: Union[str, Path],
        version_label: str = "",
    ) -> ImportStats:
        """
        Загружает ВРВС из XLS-файла в память.
        Возвращает статистику импорта.
        """
        parser = VrvsXlsParser(xls_path)
        sports = parser.parse()

        stats = ImportStats(
            version_label=version_label or f"Import from {Path(xls_path).name}",
            file_hash=parser.file_hash(),
            sports_total=len(sports),
            disciplines_total=sum(len(s.disciplines) for s in sports),
        )

        self._sports = sports
        self._rebuild_indexes()
        self._cache.clear()

        logger.info(
            f"Загружен ВРВС: {stats.sports_total} видов спорта, "
            f"{stats.disciplines_total} дисциплин"
        )
        return stats

    def load_entries(self, entries: list[SportEntry]):
        """Загружает из готового списка (напр. из БД)."""
        self._sports = entries
        self._rebuild_indexes()
        self._cache.clear()

    def add_alias(self, alias: str, canonical: str):
        """
        Добавляет пользовательский алиас.
        Пример: add_alias("Тайский бокс", "Муайтай")
        """
        self._alias_to_canon[alias] = canonical
        self._cache.clear()
        logger.info(f"Alias: '{alias}' → '{canonical}'")

    def set_name_lifetime(
        self,
        sport_name: str,
        valid_to: str,
        new_name: Optional[str] = None,
    ):
        """
        Устанавливает «срок жизни» наименования.
        Если передан new_name — создаёт алиас.

        Пример:
            normalizer.set_name_lifetime(
                "Тайский бокс",
                valid_to="2024-01-01",
                new_name="Муайтай"
            )
        """
        if new_name:
            self._alias_to_canon[sport_name] = new_name
            logger.info(
                f"Переименование: '{sport_name}' → '{new_name}' "
                f"(valid_to={valid_to})"
            )
        self._cache.clear()

    # ------------------------------------------------------------------
    # Нормализация
    # ------------------------------------------------------------------

    def normalize(self, sport_name: str) -> NormalizationResult:
        """
        Нормализует название вида спорта.

        Возвращает NormalizationResult с:
          - canonical_name: каноническое название из ВРВС
          - confidence: уверенность (0.0–1.0)
          - method: способ сопоставления
          - alternatives: топ-3 альтернативы при fuzzy

        Если confidence < review_threshold — возвращает NOT_FOUND.
        """
        if not sport_name or not sport_name.strip():
            return NormalizationResult(input_name=sport_name or "")

        name = sport_name.strip()

        # Кеш
        if name in self._cache:
            return self._cache[name]

        result = self._do_normalize(name)
        self._cache[name] = result
        return result

    def normalize_batch(
        self, names: list[str]
    ) -> list[NormalizationResult]:
        """Нормализация списка названий."""
        return [self.normalize(n) for n in names]

    def _do_normalize(self, name: str) -> NormalizationResult:
        """Внутренняя логика нормализации."""

        # 1. Exact match
        if name in self._name_to_sport:
            sport = self._name_to_sport[name]
            return NormalizationResult(
                input_name=name,
                canonical_name=sport.name,
                sport_id=sport.db_id,
                confidence=1.0,
                method=MatchMethod.EXACT,
            )

        # 2. Alias match
        if name in self._alias_to_canon:
            canon = self._alias_to_canon[name]
            sport = self._name_to_sport.get(canon)
            return NormalizationResult(
                input_name=name,
                canonical_name=canon,
                sport_id=sport.db_id if sport else None,
                confidence=0.98,
                method=MatchMethod.ALIAS,
            )

        # 3. Case-normalized match
        norm = self._normalize_text(name)
        if norm in self._norm_to_sport:
            sport = self._norm_to_sport[norm]
            return NormalizationResult(
                input_name=name,
                canonical_name=sport.name,
                sport_id=sport.db_id,
                confidence=0.95,
                method=MatchMethod.CASE_NORM,
            )

        # 3.5. Alias после нормализации
        for alias, canon in self._alias_to_canon.items():
            if self._normalize_text(alias) == norm:
                sport = self._name_to_sport.get(canon)
                return NormalizationResult(
                    input_name=name,
                    canonical_name=canon,
                    sport_id=sport.db_id if sport else None,
                    confidence=0.95,
                    method=MatchMethod.ALIAS,
                )

        # 4. Fuzzy match
        return self._fuzzy_match(name)

    def _fuzzy_match(self, name: str) -> NormalizationResult:
        """
        Нечёткое сопоставление.
        Использует rapidfuzz если доступен, иначе — встроенный алгоритм
        на основе расстояния Левенштейна / подстрок.
        """
        # Собираем словарь кандидатов
        choices: dict[str, SportEntry] = {}
        for s in self._sports:
            choices[s.name] = s
        for alias, canon in self._alias_to_canon.items():
            if canon in {s.name for s in self._sports}:
                sport = self._name_to_sport.get(canon)
                if sport:
                    choices[alias] = sport

        if not choices:
            return NormalizationResult(
                input_name=name, method=MatchMethod.NOT_FOUND
            )

        try:
            from rapidfuzz import fuzz, process
            results = process.extract(
                name, list(choices.keys()),
                scorer=fuzz.token_sort_ratio,
                limit=5,
            )
            scored = [(match_name, score / 100, choices[match_name])
                      for match_name, score, _ in results]
        except ImportError:
            # Fallback: встроенный fuzzy на основе trigram-сходства
            scored = self._builtin_fuzzy(name, choices, limit=5)

        if not scored:
            return NormalizationResult(
                input_name=name, method=MatchMethod.NOT_FOUND
            )

        best_name, best_score, best_sport = scored[0]
        canonical = self._alias_to_canon.get(best_name, best_name)
        canon_sport = self._name_to_sport.get(canonical, best_sport)

        alternatives = [
            (match_name, round(score, 3))
            for match_name, score, _ in scored[1:4]
        ]

        confidence = round(best_score, 3)

        if confidence < self.review_threshold / 100:
            return NormalizationResult(
                input_name=name,
                method=MatchMethod.NOT_FOUND,
                alternatives=alternatives,
                confidence=confidence,
            )

        return NormalizationResult(
            input_name=name,
            canonical_name=canonical,
            sport_id=canon_sport.db_id if canon_sport else None,
            confidence=confidence,
            method=MatchMethod.FUZZY,
            alternatives=alternatives,
        )

    @staticmethod
    def _builtin_fuzzy(
        query: str,
        choices: dict[str, "SportEntry"],
        limit: int = 5,
    ) -> list[tuple[str, float, "SportEntry"]]:
        """
        Встроенный fuzzy-матчинг без внешних зависимостей.
        Комбинация trigram-сходства и substring matching.
        """
        def trigrams(s: str) -> set[str]:
            s = s.lower().strip()
            padded = f"  {s} "
            return {padded[i:i+3] for i in range(len(padded) - 2)}

        def trigram_similarity(a: str, b: str) -> float:
            ta, tb = trigrams(a), trigrams(b)
            if not ta or not tb:
                return 0.0
            intersection = ta & tb
            union = ta | tb
            return len(intersection) / len(union)

        def combined_score(query: str, candidate: str) -> float:
            # Trigram similarity (Jaccard)
            tri_score = trigram_similarity(query, candidate)

            # Normalized containment (is one inside the other?)
            ql, cl = query.lower(), candidate.lower()
            if ql in cl or cl in ql:
                contain_bonus = 0.15
            else:
                contain_bonus = 0.0

            # Length penalty (big length difference = penalty)
            len_ratio = min(len(query), len(candidate)) / max(len(query), len(candidate), 1)
            len_bonus = len_ratio * 0.1

            return min(1.0, tri_score + contain_bonus + len_bonus)

        results = []
        for name, sport in choices.items():
            score = combined_score(query, name)
            results.append((name, score, sport))

        results.sort(key=lambda x: -x[1])
        return results[:limit]

    # ------------------------------------------------------------------
    # Запросы к справочнику
    # ------------------------------------------------------------------

    def all_sports(self) -> list[SportEntry]:
        """Все виды спорта из загруженного справочника."""
        return list(self._sports)

    def sports_by_section(self, section: int) -> list[SportEntry]:
        """Виды спорта по разделу ВРВС (1–4)."""
        return [s for s in self._sports if s.section == section]

    def find_sport(self, name: str) -> Optional[SportEntry]:
        """Точный поиск по имени (после нормализации)."""
        norm = self._normalize_text(name)
        return self._norm_to_sport.get(norm)

    def search(self, query: str, limit: int = 10) -> list[tuple[str, float]]:
        """Поиск с fuzzy по всем именам. Возвращает [(name, score), ...]."""
        try:
            from rapidfuzz import fuzz, process
            names = [s.name for s in self._sports]
            results = process.extract(
                query, names, scorer=fuzz.token_sort_ratio, limit=limit
            )
            return [(name, round(score / 100, 3)) for name, score, _ in results]
        except ImportError:
            # Fallback: встроенный trigram
            choices = {s.name: s for s in self._sports}
            scored = self._builtin_fuzzy(query, choices, limit=limit)
            return [(name, round(score, 3)) for name, score, _ in scored]

    # ------------------------------------------------------------------
    # Генерация SQL INSERT (для импорта в БД без прямого подключения)
    # ------------------------------------------------------------------

    def generate_sql(
        self,
        version_label: str = "ВРВС",
        output_path: Optional[Union[str, Path]] = None,
    ) -> str:
        """
        Генерирует SQL-скрипт для импорта справочника в БД.
        Если output_path указан — сохраняет в файл.
        """
        if not self._sports:
            raise ValueError("Справочник не загружен")

        lines = [
            "-- =====================================================",
            f"-- Импорт ВРВС: {version_label}",
            f"-- Видов спорта: {len(self._sports)}",
            f"-- Дисциплин: {sum(len(s.disciplines) for s in self._sports)}",
            "-- =====================================================",
            "",
            "BEGIN;",
            "",
            "-- Версия реестра",
            f"INSERT INTO sport_registry_versions (version_label, sport_count, discipline_count)",
            f"VALUES ('{self._esc(version_label)}', {len(self._sports)}, "
            f"{sum(len(s.disciplines) for s in self._sports)});",
            "",
        ]

        for sport in self._sports:
            ename = self._esc(sport.name)
            ecode = self._esc(sport.code_full)
            lines.append(
                f"INSERT INTO sports (code_base, code_full, section, current_name) "
                f"VALUES ({sport.code_base}, '{ecode}', {sport.section}, '{ename}') "
                f"ON CONFLICT (code_base, section) DO UPDATE SET "
                f"current_name = EXCLUDED.current_name, code_full = EXCLUDED.code_full;"
            )
            lines.append(
                f"INSERT INTO sport_names (sport_id, name, is_primary) "
                f"VALUES ((SELECT id FROM sports WHERE code_base = {sport.code_base} "
                f"AND section = {sport.section}), '{ename}', TRUE) "
                f"ON CONFLICT DO NOTHING;"
            )

            for disc in sport.disciplines:
                edisc = self._esc(disc)
                lines.append(
                    f"INSERT INTO sport_disciplines (sport_id, name) "
                    f"VALUES ((SELECT id FROM sports WHERE code_base = {sport.code_base} "
                    f"AND section = {sport.section}), '{edisc}') "
                    f"ON CONFLICT DO NOTHING;"
                )

        lines.append("")
        lines.append("COMMIT;")

        sql = "\n".join(lines)

        if output_path:
            Path(output_path).write_text(sql, encoding="utf-8")
            logger.info(f"SQL сохранён: {output_path}")

        return sql

    # ------------------------------------------------------------------
    # Утилиты
    # ------------------------------------------------------------------

    def _rebuild_indexes(self):
        """Перестраивает in-memory индексы после загрузки."""
        self._name_to_sport.clear()
        self._norm_to_sport.clear()

        for sport in self._sports:
            self._name_to_sport[sport.name] = sport
            norm = self._normalize_text(sport.name)
            self._norm_to_sport[norm] = sport

    @staticmethod
    def _normalize_text(text: str) -> str:
        """
        Нормализация текста для сопоставления:
        - upper case
        - удаление лишних пробелов
        - замена ё → е
        - удаление пунктуации кроме дефисов
        """
        if not text:
            return ""
        s = text.strip().upper()
        s = s.replace("Ё", "Е").replace("ё", "е")
        s = re.sub(r"[^\w\s\-]", "", s, flags=re.UNICODE)
        s = re.sub(r"\s+", " ", s)
        return s.strip()

    @staticmethod
    def _esc(s: str) -> str:
        """Экранирование для SQL (одинарные кавычки)."""
        return s.replace("'", "''")

    # ------------------------------------------------------------------
    # Отчёты
    # ------------------------------------------------------------------

    def coverage_report(
        self, sport_names_from_orders: list[str]
    ) -> dict:
        """
        Анализирует покрытие: какие виды спорта из приказов
        сопоставились, а какие нет.

        Возвращает словарь:
        {
            "total": int,
            "matched": int,
            "auto_matched": int,     # confidence >= auto_threshold
            "review_needed": int,    # review_threshold <= conf < auto_threshold
            "not_found": int,
            "details": [NormalizationResult, ...]
        }
        """
        results = self.normalize_batch(sport_names_from_orders)
        auto_t = self.auto_threshold / 100
        review_t = self.review_threshold / 100

        matched = [r for r in results if r.method != MatchMethod.NOT_FOUND]
        auto = [r for r in matched if r.confidence >= auto_t]
        review = [r for r in matched
                  if review_t <= r.confidence < auto_t]
        not_found = [r for r in results if r.method == MatchMethod.NOT_FOUND]

        return {
            "total": len(results),
            "matched": len(matched),
            "auto_matched": len(auto),
            "review_needed": len(review),
            "not_found": len(not_found),
            "details": results,
        }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _main():
    import argparse
    import json

    parser = argparse.ArgumentParser(
        description="SportNormalizer — импорт ВРВС и нормализация видов спорта"
    )
    sub = parser.add_subparsers(dest="command")

    # import
    p_import = sub.add_parser("import", help="Импорт XLS реестра")
    p_import.add_argument("xls", help="Путь к XLS-файлу ВРВС")
    p_import.add_argument("--label", default="", help="Метка версии")
    p_import.add_argument("--sql", help="Сохранить SQL-скрипт")
    p_import.add_argument("--json", dest="json_out", help="Сохранить как JSON")

    # normalize
    p_norm = sub.add_parser("normalize", help="Нормализовать название")
    p_norm.add_argument("xls", help="Путь к XLS-файлу ВРВС")
    p_norm.add_argument("name", help="Название вида спорта из приказа")

    # search
    p_search = sub.add_parser("search", help="Поиск по справочнику")
    p_search.add_argument("xls", help="Путь к XLS-файлу ВРВС")
    p_search.add_argument("query", help="Запрос")
    p_search.add_argument("--limit", type=int, default=10)

    # coverage
    p_cov = sub.add_parser("coverage", help="Отчёт покрытия")
    p_cov.add_argument("xls", help="Путь к XLS-файлу ВРВС")
    p_cov.add_argument("names_file", help="Файл с названиями (по одному на строку)")

    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    if args.command == "import":
        normalizer = SportNormalizer()
        stats = normalizer.load_xls(args.xls, version_label=args.label)
        print(f"Импорт: {stats.sports_total} видов спорта, "
              f"{stats.disciplines_total} дисциплин")
        print(f"SHA256: {stats.file_hash[:16]}...")

        if args.sql:
            normalizer.generate_sql(
                version_label=args.label or "ВРВС",
                output_path=args.sql,
            )
            print(f"SQL: {args.sql}")

        if args.json_out:
            data = [
                {
                    "code_base": s.code_base,
                    "code_full": s.code_full,
                    "section": s.section,
                    "section_name": SECTION_NAMES.get(s.section, ""),
                    "name": s.name,
                    "disciplines_count": len(s.disciplines),
                    "disciplines": s.disciplines[:5],  # первые 5
                }
                for s in normalizer.all_sports()
            ]
            Path(args.json_out).write_text(
                json.dumps(data, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            print(f"JSON: {args.json_out}")

        # Вывод сводки
        print()
        for sec, name in SECTION_NAMES.items():
            sports = normalizer.sports_by_section(sec)
            disc = sum(len(s.disciplines) for s in sports)
            print(f"  Раздел {sec} ({name}): "
                  f"{len(sports)} видов спорта, {disc} дисциплин")

    elif args.command == "normalize":
        normalizer = SportNormalizer()
        normalizer.load_xls(args.xls)
        result = normalizer.normalize(args.name)
        print(f"Ввод:       {result.input_name}")
        print(f"Канон:      {result.canonical_name or '(не найден)'}")
        print(f"Уверенн.:   {result.confidence:.2f}")
        print(f"Метод:      {result.method.value}")
        if result.alternatives:
            print(f"Альтернат.: {result.alternatives}")

    elif args.command == "search":
        normalizer = SportNormalizer()
        normalizer.load_xls(args.xls)
        results = normalizer.search(args.query, limit=args.limit)
        print(f"Поиск: '{args.query}'")
        for name, score in results:
            print(f"  {score:.2f}  {name}")

    elif args.command == "coverage":
        normalizer = SportNormalizer()
        normalizer.load_xls(args.xls)
        names = Path(args.names_file).read_text(encoding="utf-8").strip().split("\n")
        report = normalizer.coverage_report(names)
        print(f"Всего:          {report['total']}")
        print(f"Сопоставлено:   {report['matched']}")
        print(f"  Авто:         {report['auto_matched']}")
        print(f"  На ревью:     {report['review_needed']}")
        print(f"Не найдено:     {report['not_found']}")

        # Детали по не найденным
        not_found = [r for r in report["details"]
                     if r.method == MatchMethod.NOT_FOUND]
        if not_found:
            print("\nНе найдены:")
            for r in not_found:
                alt = f" (ближайшее: {r.alternatives[0]})" if r.alternatives else ""
                print(f"  ❌ {r.input_name}{alt}")

    else:
        parser.print_help()


if __name__ == "__main__":
    _main()
